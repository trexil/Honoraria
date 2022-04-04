from pywebio import *
from pywebio.utils import pyinstaller_datas
from pywebio.input import *
from pywebio.output import *
from pywebio.session import *
from pywebio.pin import *
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import argparse
import io
import os
import json
import operator

professor=[]
with open(os.getcwd()+ '\\config.txt', 'r') as f:
    for names in f.readlines():
        if names.strip():
            professor.append(names.rstrip('\n'))
        elif not names.strip():
            continue

jsonpath = os.getcwd() + '\\Files\\Json\\' #directory of JSON file
excelpath = os.getcwd() + '\\Files\\Excel\\' #directory of backup excelfile



def startupCheck():
    Thesis = {
        'Thesis': [
        ],
        'Report': [
        ]}
    if os.path.isfile(jsonpath + '\\Thesis.json'):
        pass
    else:
        with open(jsonpath+'Thesis.json', 'w') as outfile:
            json.dump(Thesis, outfile, indent=4)


def InitialJSon(filename, professor):
    professor.sort()
    with open(filename) as f:
        list = json.load(f)
        tempR = list['Report']
        for name in professor:
            x = {
                'prof': name,
                'advisor': 0,
                'panel': 0
            }
            tempR.append(x)
            with open(filename, 'w') as g:
                json.dump(list, g, indent=4)

'''
def startupCheck(professors):
    num=len(professors)
    profs = professors.sort()
    Thesis = {
        'Thesis': [
        ],
        'Report': [
            {
                'prof': 'Trex',
                'advisor': 0,
                'panel': 0
            },
            {
                'prof': 'Caya, Meo Vincent',
                'advisor': 0,
                'panel': 0
            },
            {
                'prof': 'Hortinela, Carlos IV',
                'advisor': 0,
                'panel': 0
            },
            {
                'prof': 'Linsangan, Noel',
                'advisor': 0,
                'panel': 0
            },
            {
                'prof': 'Manlises, Cyrel',
                'advisor': 0,
                'panel': 0
            },
            {
                'prof': 'Maramba, Rafael',
                'advisor': 0,
                'panel': 0
            },
            {
                'prof': 'Pellegrino, Rosemarie',
                'advisor': 0,
                'panel': 0
            },
            {
                'prof': 'Torres, Jumelyn',
                'advisor': 0,
                'panel': 0
            },
            {
                'prof': 'Villaverde, Jocelyn',
                'advisor': 0,
                'panel': 0
            },
            {
                'prof': 'Yumang, Analyn',
                'advisor': 0,
                'panel': 0
            },
        ]
    }
    if os.path.isfile(jsonpath + '\\Thesis.json'):
        pass
    else:
        with open (jsonpath+'Thesis.json', 'w') as outfile:
            json.dump(Thesis, outfile, indent=4)
'''

jsonfile = jsonpath + 'Thesis.json'
reportfile = jsonpath + 'Report.json'

def write_json(filename):
    data = datagather(professor)
    if data == 0:
        main_menu()
    else:
        x = {
            'title': data[0],
            'advisor': data[1],
            'panels': data[2],
            'payor': data[3],
            'refnum': data[4],
            'paid': 'Not yet paid'
        }
        with open(filename) as f:
            list = json.load(f)
            temp = list['Thesis']
            temp.append(x)
            temp.sort(key=operator.itemgetter('title'))
        with open(filename,'w') as f:
            json.dump(list, f, indent=4)



def main_menu():
    with use_scope('menu_scope'):
        data = input_group('Honorarium',[
            actions('Select:',[
                {'label': 'Browse Data', 'value': 'Browse Data'},
                {'label': 'Add New Thesis', 'value': 'Add New Thesis'},
                {'label': 'Report', 'value': 'Report'},
                {'label': 'Reset', 'value': 'Reset', 'color': 'danger'},
                {'label': 'Close', 'value': 'Close', 'color': 'warning'}
            ], name='response'),
            ])
    return data['response']

def index():
    response = main_menu()
    while response != 'Close':
        if response == 'Browse Data':
            browse(jsonfile)
        elif response == 'Add New Thesis':
            write_json(jsonfile)
        elif response == 'Report':
            reportMenu(jsonfile, professor)
        elif response == 'Reset':
            reset(jsonfile)
        response = main_menu()

def reset(filename):
    Thesis = {
        'Thesis': [
        ],
        'Report': [
        ]}
    data = input_group("Are you sure you want to erase all data?", [
        actions(' ', [
            {'label': 'Yes', 'value': 'Yes', 'color': 'danger'},
            {'label': 'No', 'value': 'No'},
        ], name='action', ),
    ])
    if data['action'] == 'Yes':
        with open(filename, 'w') as outfile:
            json.dump(Thesis, outfile, indent=4)
        index()
    else:
        index()



def browse(filename):
    with open(filename) as f:
        list = json.load(f)
        temp = list['Thesis']
        x = 0
        with popup('Thesis'):
            while x < len(temp):
                with put_collapse(temp[x]['title']):
                    put_table([
                        ['Title', temp[x]['title']],
                        ['Advisor', put_text(';'.join(temp[x]['advisor']))],
                        ['Panels', put_text(';'.join(temp[x]['panels']))],
                        ['Payor', temp[x]['payor']],
                        ['Reference Number', temp[x]['refnum']],
                        ['Payment Status', temp[x]['paid']]
                    ])
                    Title =temp[x]['title']
                    Advisor = temp[x]['advisor']
                    Panels = temp[x]['panels']
                    Payor = temp[x]['payor']
                    Ref = temp[x]['refnum']
                    Paid = temp[x]['paid']
                    column = len(Advisor) + len(Panels) + 4
                    fname = Title + ".xlsx"
                    honorariumReport(Title, Advisor, Panels, Payor, Ref, column, fname)
                    paidbutton(Title, jsonfile)
                    deleteEntry(Title, jsonfile)
                x += 1
            put_button('Back', onclick=lambda: close_popup())


def paidFunction(title, filename):
    with open(filename) as f:
        list = json.load(f)
        temp = list['Thesis']
        for data in temp:
            if data['title'] == title:
                if data['paid']== 'Not yet paid':
                    data['paid'] = 'Paid'
                else:
                    data['paid'] = 'Not yet paid'
                with open(filename, 'w') as f:
                    json.dump(list, f, indent=4)
                browse(jsonfile)


def paidbutton(title, filename):
    put_button('Change payment status', lambda: paidFunction(title, jsonfile), color='warning', )

def deleteFunction(title, filename):
    with open(filename) as f:
        list = json.load(f)
        temp = list['Thesis']
        x=0
        for data in temp:
            if data['title'] == title:
                temp.pop(x)
                with open(filename, 'w') as f:
                    json.dump(list, f, indent=4)
                browse(jsonfile)
            else:
                x+=1

def deleteEntry(title, filename):
    put_button('Delete', lambda: deleteFunction(title, jsonfile), color='danger',)

def reportMenu(filename,professor):
    fstring = [['Professor', 'Advisor', 'Panel', 'Total(Pesos)']]
    professor.sort()
    with open(filename) as f:
        list = json.load(f)
        tempR = list['Report']
        tempR.clear()
        for name in professor:
            x = {
                'prof': name,
                'advisor': 0,
                'panel': 0
            }
            tempR.append(x)
            with open(filename, 'w') as g:
                json.dump(list, g, indent=4)
    with open(filename) as f:
        list = json.load(f)
        tempT = list['Thesis']
        tempR = list['Report']
        for data in tempT:
            if data['paid'] == 'Not yet paid':
                for advisor in data['advisor']:
                    for dataR in tempR:
                        if advisor==dataR['prof']:
                            dataR['advisor']+=1

                for panel in data['panels']:
                    for dataR in tempR:
                        if panel==dataR['prof']:
                            dataR['panel']+=1
        with open(filename, 'w') as f:
            json.dump(list, f, indent=4)
        for x in range(len(professor)):
            dataR = [tempR[x]['prof'], tempR[x]['advisor'], tempR[x]['panel'],(tempR[x]['advisor'] * 3000 + tempR[x]['panel'] * 1500)]
            fstring.append(dataR)

        with popup('Report'):
            put_table(
                fstring
            )
            reportexcel(professor)
            put_button('Back', onclick=lambda: close_popup(), color='danger')

def reportexcel(professor):
    with open(jsonfile) as f:
        list = json.load(f)
        tempT = list['Thesis']
        tempR = list['Report']
        '''for data in tempT:
            if data['paid'] == 'Not yet paid':
                for advisor in data['advisor']:
                    for dataR in tempR:
                        if advisor==dataR['prof']:
                            dataR['advisor']+=1

                for panel in data['panels']:
                    for dataR in tempR:
                        if panel==dataR['prof']:
                            dataR['panel']+=1
'''
        column=len(professor) + 3
        Y = ["A", 'B', 'C', 'D']
        wb = Workbook()
        ws = wb.active
        fname = 'Honorarium Shared Report' + '.xlsx'

        ws.merge_cells('A1:A2')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions['A'].width = 20
        ws['A1'] = "FACULTY NAME"
        for x in range(len(professor)):
            ws['A'+ str(x+3)] = tempR[x]['prof']


        ws.merge_cells('B1:B2')
        counterB = 1
        while counterB != column:
            ws['B' + str(counterB)].alignment = Alignment(vertical='center', horizontal='center', wrapText=True)
            counterB += 1
        ws.column_dimensions['B'].width = 10
        ws['B1'] = "ADVISOR"
        cB = 3
        for data in tempR:
            ws['B'+str(cB)] = data['advisor']
            cB+=1

        ws.merge_cells('C1:C2')
        counterC = 1
        while counterC != column:
            ws['C' + str(counterC)].alignment = Alignment(vertical='center', horizontal='center', wrapText=True)
            counterC += 1
        ws.column_dimensions['C'].width = 10
        ws['C1'] = "PANEL"
        cC = 3
        for data in tempR:
            ws['C' + str(cC)] = data['panel']
            cC += 1

        ws.merge_cells('D1:D2')
        counterD = 1
        while counterD < column:
            ws['D' + str(counterD)].alignment = Alignment(vertical='center', horizontal='center')
            counterD += 1
        ws.column_dimensions['D'].width = len("TOTAL(Pesos)") + 5
        ws['D1'] = "TOTAL(Pesos)"
        cD = 3
        for data in tempR:
            ws['D' + str(cD)] = 'PHP ' + str(data['advisor']*3000 + data['panel']*1500)
            cD += 1


        for C in Y:
            for R in range(column-1):
                ws[str(C) + str(R + 1)].font = Font(name="Arial")
                ws[str(C) + str(R + 1)].border = Border(left=Side(border_style='thin', color='00000000'),
                                                        right=Side(border_style='thin', color='00000000'),
                                                        top=Side(border_style='thin', color='00000000'),
                                                        bottom=Side(border_style='thin', color='00000000'))

        wb.save(excelpath + fname)
        in_file = open(excelpath + fname, 'rb')
        data = in_file.read()
        in_file.close()
        put_buttons(['Download Report', 'Mark All as Paid'], onclick=[lambda: download(fname, data), lambda: markpaid(jsonfile)])

def markpaid(filename):
    with open(filename) as f:
        list = json.load(f)
        temp = list['Thesis']
        for data in temp:
            if data['paid'] == 'Not yet paid':
                data['paid'] = 'Paid'
                with open(filename, 'w') as f:
                    json.dump(list, f, indent=4)
    reportMenu(jsonfile, professor)

def datagather(professors):
    data = input_group("Thesis/Design Information", [
        input('Title of Thesis/Design', name='title'),
        checkbox('Advisor', options=professors, name='advisor'),
        checkbox('Panels', options=professors, name='panels'),
        input("Payor Name", name='payor'),
        input("Reference Number", name='refnum'),
        actions('actions', [
            {'label': 'Save', 'value': 'save'},
            {'label': 'Reset', 'type': 'reset', 'color': 'warning'},
            {'label': 'Back', 'value': 'back', 'color':'danger'},
        ], name='action', help_text='actions'),
    ])
    if data['action'] == 'back':
        return 0
    elif not data['title']:
        popup('Error', [
            put_text('Missing Title')])
        datagather(professor)
    elif not data['advisor']:
        popup('Error', [
            put_text('Missing Advisor')])
        datagather(professor)
    elif not data['panels']:
        popup('Error', [
            put_text('Missing Panels')])
        datagather(professor)
    elif not data['payor']:
        popup('Error', [
            put_text('Missing Payor')])
        datagather(professor)
    elif not data['refnum']:
        popup('Error', [
            put_text('Missing Reference Number')])
        datagather(professor)
    elif data['action']=='save':
        return(data['title'],data['advisor'],data['panels'],data['payor'],data['refnum'])


def honorariumReport(Title,Advisor,Panels,Payor,Ref,column,fname):
    Y = ["A", 'B', 'C', 'D', 'E', 'F']
    wb = Workbook()
    ws = wb.active
    ws.title = Title

    ws.merge_cells(start_row=1, start_column=1, end_row=column, end_column=1)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['A'].width = 4
    ws['A1'] = 1

    ws.merge_cells('B1:B2')
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['B'].width = 20
    ws['B1'] = "FACULTY NAME"
    x = 3
    for adv in Advisor:
        ws['B' + str(x)] = adv
        ws['C' + str(x)] = "Advisor"
        ws['E' + str(x)] = "PHP 3,000"
        ws['F' + str(x)] = "ATM"
        x += 1
    y = len(Advisor) + 4
    for pan in Panels:
        ws['B' + str(y)] = pan
        ws['C' + str(y)] = "Panel"
        ws['E' + str(y)] = "PHP 1,500"
        ws['F' + str(y)] = "ATM"
        y += 1

    ws.merge_cells('C1:C2')
    counterC = 1
    while counterC != column:
        ws['C' + str(counterC)].alignment = Alignment(vertical='center', horizontal='center', wrapText=True)
        counterC += 1
    ws.column_dimensions['C'].width = 10
    ws['C1'] = "ADVISOR or PANEL"

    ws.merge_cells('D1:D2')
    counterD = 1
    while counterD <= column:
        ws['D' + str(counterD)].alignment = Alignment(vertical='center', horizontal='center', wrapText=True)
        counterD += 1
    ws.merge_cells(start_row=3, start_column=4, end_row=column - 2, end_column=4)
    ws.column_dimensions['D'].width = 30
    ws['D1'] = 'TITLE OF THESIS/DESIGN'
    ws['D3'] = Title
    ws['D' + str(column - 1)] = Payor
    ws['D' + str(column)] = Ref

    ws.merge_cells('E1:E2')
    counterE = 1
    while counterE < column:
        ws['E' + str(counterE)].alignment = Alignment(vertical='center', horizontal='center')
        counterE += 1
    ws.column_dimensions['E'].width = len("HONORARIUM") + 5
    ws['E1'] = "HONORARIUM"

    ws.merge_cells('F1:F2')
    counterF = 1
    while counterF < column:
        ws['F' + str(counterF)].alignment = Alignment(vertical='center', horizontal='center', wrapText=True)
        counterF += 1
    ws.column_dimensions['F'].width = 12
    ws['F1'] = "MODE OF PAYMENT"

    for C in Y:
        for R in range(column):
            ws[str(C) + str(R + 1)].font = Font(name="Arial")
            ws[str(C) + str(R + 1)].border = Border(left=Side(border_style='thin', color='00000000'),
                                                    right=Side(border_style='thin', color='00000000'),
                                                    top=Side(border_style='thin', color='00000000'),
                                                    bottom=Side(border_style='thin', color='00000000'))

    wb.save(excelpath + fname)
    in_file = open(excelpath + fname, 'rb')
    data = in_file.read()
    in_file.close()
    put_button('Download Excel File', lambda: download(fname, data))

def main():
    startupCheck()
    InitialJSon(jsonfile, professor)
    index()

#main()


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("-p", "--port", type=int, default=8080)
    args = parser.parse_args()
    start_server(main, debug=True, port=args.port)

